import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_paths = [
    r"E:\project\word_card\word_card_master_300.xlsx",
    r"E:\project\word_card\word_assets_all.xlsx"
]

for ep in excel_paths:
    if not os.path.exists(ep): continue
    wb = openpyxl.load_workbook(ep)
    if "전체_자모_타일_목록" in wb.sheetnames:
        ws = wb["전체_자모_타일_목록"]
        for row in range(3, ws.max_row + 1):
            char_val = str(ws.cell(row=row, column=1).value or "").strip()
            if char_val in ["ㄱ", "ㅡ", "ㅜ"]:
                ws.cell(row=row, column=2).value = "🌟 희귀 자모 (Rare, 낮은 확률)"
                ws.cell(row=row, column=5).value = "희귀 드롭 풀 (가중치 20, 일반의 1/5)"
    wb.save(ep)
    print(f"Updated rarity in: {ep}")
